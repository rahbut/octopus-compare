#!/usr/bin/env python3
"""
Ofgem Price Cap Data Builder
=============================
Downloads quarterly Ofgem price cap files and extracts regional direct debit
unit rates and standing charges for electricity and gas, outputting a single
JSON lookup file for use in the Octopus Energy tariff comparison tool.

Usage:
    python build_ofgem_data.py

Output:
    ofgem-price-cap.json

Requirements:
    pip install requests openpyxl

Quarterly update process:
    1. Go to https://www.ofgem.gov.uk/energy-regulation/domestic-and-non-domestic/energy-pricing-rules/energy-price-cap/energy-price-cap-default-tariff-levels
    2. Find the new quarter's section
    3. For quarters FROM April 2024 onwards:
         Copy the URL of "Final levelised cap rates model (Annex 9)"
         Add a new entry to ANNEX9_QUARTERS below
    4. For any quarter BEFORE April 2024 (unlikely to need updating):
         Copy the URL of "Energy price cap levels: Pre-levelised rates model"
         Add to PRE_LEVELISED_QUARTERS below
    5. Run this script — it will regenerate the full JSON file

Notes:
    - All rates are in pence (inc. VAT), as shown on consumer bills
    - Standing charges are in pence per day
    - Unit rates are in pence per kWh
    - Payment method: Direct Debit throughout (most common for Octopus customers)
    - Northern Ireland is NOT covered by the Ofgem cap
"""

import json
import os
import re
import sys
from datetime import date

try:
    import requests
    import openpyxl
except ImportError:
    print("Missing dependencies. Run: pip install requests openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Region definitions
# The 14 Grid Supply Point regions covered by the Ofgem cap.
# These match the GSP codes returned by the Octopus API.
# ---------------------------------------------------------------------------
REGIONS = {
    "_A": "Eastern England",
    "_B": "East Midlands",
    "_C": "London",
    "_D": "Merseyside and North Wales",
    "_E": "West Midlands",
    "_F": "North East England",
    "_G": "North West England",
    "_H": "Southern England",
    "_J": "South East England",
    "_K": "South Western England",
    "_L": "South Wales",
    "_M": "Yorkshire",
    "_N": "South Scotland",
    "_P": "North Scotland",
}

# Ofgem uses slightly different region name spellings across file versions.
# This maps known variants to our canonical GSP codes.
REGION_NAME_TO_GSP = {
    # Canonical names
    "eastern england": "_A",
    "east midlands": "_B",
    "london": "_C",
    "merseyside and north wales": "_D",
    "north wales and mersey": "_D",       # alternate spelling used in some files
    "west midlands": "_E",
    "north east england": "_F",
    "north east": "_F",
    "north west england": "_G",
    "north west": "_G",
    "southern england": "_H",
    "south east england": "_J",
    "south east": "_J",
    "south western england": "_K",
    "south west": "_K",
    "south wales": "_L",
    "yorkshire": "_M",
    "south scotland": "_N",
    "north scotland": "_P",
    "scotland (north)": "_P",
    "scotland (south)": "_N",
    # Names used in the 1a cap/DTC sheets (both pre-levelised and Annex 9)
    "north west": "_G",
    "northern": "_F",           # Ofgem's "Northern" = NE England (Northern Powergrid)
    "northern scotland": "_P",
    "southern": "_H",
    "southern scotland": "_N",
    "n wales and mersey": "_D",
    "eastern": "_A",
    "midlands": "_E",           # West Midlands (Western Power Distribution Midlands)
    "southern western": "_K",
}


# ---------------------------------------------------------------------------
# File catalogue
#
# PRE_LEVELISED_QUARTERS: Oct 2021 – Jan 2024
#   Source: "Energy price cap levels: Pre-levelised rates model" XLSX
#   These files contain a sheet called something like "Cap table" or similar,
#   with regional rates for DD customers buried within.
#   We target the sheet containing "unit rate" and "standing charge" rows.
#
# ANNEX9_QUARTERS: Apr 2024 onwards
#   Source: "Final levelised cap rates model (Annex 9)" XLSX
#   These are cleaner — a dedicated sheet with one row per region showing
#   final consumer-facing DD rates for electricity and gas.
# ---------------------------------------------------------------------------

PRE_LEVELISED_QUARTERS = [
    {
        "period_from": "2021-10-01",
        "period_to":   "2022-03-31",
        "label":       "Q4 2021 / Q1 2022",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2021-08/Default_tariff_cap_level_v1.9.xlsx",
    },
    {
        "period_from": "2022-04-01",
        "period_to":   "2022-09-30",
        "label":       "Q2/Q3 2022",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2022-02/Default_tarif_cap_level_v1.10.xlsx",
    },
    {
        "period_from": "2022-10-01",
        "period_to":   "2022-12-31",
        "label":       "Q4 2022",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2022-08/Default_tariff_cap_level_v1.13.xlsx",
    },
    {
        "period_from": "2023-01-01",
        "period_to":   "2023-03-31",
        "label":       "Q1 2023",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2022-11/Default_tariff_cap_level_v1.14.xlsx",
    },
    {
        "period_from": "2023-04-01",
        "period_to":   "2023-06-30",
        "label":       "Q2 2023",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2023-02/Default_tariff_cap_level_v1.15.xlsx",
    },
    {
        "period_from": "2023-07-01",
        "period_to":   "2023-09-30",
        "label":       "Q3 2023",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2023-05/Default_tariff_cap_level_v1.18_0.xlsx",
    },
    {
        "period_from": "2023-10-01",
        "period_to":   "2023-12-31",
        "label":       "Q4 2023",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2023-08/Default_tariff_cap_level_v1.19.xlsx",
    },
    {
        "period_from": "2024-01-01",
        "period_to":   "2024-03-31",
        "label":       "Q1 2024",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2023-11/Default_tariff_cap_level_v1.20.xlsx",
    },
]

ANNEX9_QUARTERS = [
    {
        "period_from": "2024-04-01",
        "period_to":   "2024-06-30",
        "label":       "Q2 2024",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2024-03/Annex_9_-_Levelisation%20allowance%20methodology%20and%20levelised%20cap%20levels_v1.1.xlsx",
    },
    {
        "period_from": "2024-07-01",
        "period_to":   "2024-09-30",
        "label":       "Q3 2024",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2024-05/Annex%209%20-%20Levelisation%20allowance%20methodology%20and%20levelised%20cap%20levels%20v1.2.xlsx",
    },
    {
        "period_from": "2024-10-01",
        "period_to":   "2024-12-31",
        "label":       "Q4 2024",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2024-08/Annex_9_-_Levelisation_allowance_methodology_and_levelised_cap_levels_v1.3.xlsx",
    },
    {
        "period_from": "2025-01-01",
        "period_to":   "2025-03-31",
        "label":       "Q1 2025",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2024-11/Annex_9_-_Levelisation_allowance_methodology_and_levelised_cap_levels_v1.4.xlsx",
    },
    {
        "period_from": "2025-04-01",
        "period_to":   "2025-06-30",
        "label":       "Q2 2025",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2025-02/Annex-9-Levelisation-allowance-methodology-and-levelised-cap-levels-v1.5.xlsx",
    },
    {
        "period_from": "2025-07-01",
        "period_to":   "2025-09-30",
        "label":       "Q3 2025",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2025-05/Final%20levelised%20cap%20rates%20model%20%28Annex%209%29%201%20July%20to%2030%20September%202025_0.xlsx",
    },
    {
        "period_from": "2025-10-01",
        "period_to":   "2025-12-31",
        "label":       "Q4 2025",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2025-08/levelisation-allowance-methodology-and-levelised-cap-levels-annex-9.xlsx",
    },
    {
        "period_from": "2026-01-01",
        "period_to":   "2026-03-31",
        "label":       "Q1 2026",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2025-11/Annex-9-Levelisation-allowance-methodology-and-levelised-cap-levels-v1.8.xlsx",
    },
    {
        "period_from": "2026-04-01",
        "period_to":   "2026-06-30",
        "label":       "Q2 2026",
        "url": "https://www.ofgem.gov.uk/sites/default/files/2026-02/Annex-9-Levelisation-allowance-methodology-and-levelised-cap-levels-v1.9.xlsx",
    },
]


# ---------------------------------------------------------------------------
# Download helper
# ---------------------------------------------------------------------------

CACHE_DIR = ".ofgem_cache"

def download_file(url: str, label: str) -> str:
    """Download a file to the local cache, returning the local path."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    filename = re.sub(r"[^\w.]", "_", url.split("/")[-1])
    local_path = os.path.join(CACHE_DIR, filename)

    if os.path.exists(local_path):
        print(f"  [cached] {label}")
        return local_path

    print(f"  [downloading] {label} ...")
    headers = {"User-Agent": "Mozilla/5.0 (compatible; ofgem-data-builder/1.0)"}
    response = requests.get(url, headers=headers, timeout=60)
    response.raise_for_status()
    with open(local_path, "wb") as f:
        f.write(response.content)
    print(f"  [saved] {local_path} ({len(response.content) / 1024:.0f} KB)")
    return local_path


# ---------------------------------------------------------------------------
# Sheet inspection helper
# ---------------------------------------------------------------------------

def find_sheet(wb: openpyxl.Workbook, candidates: list[str]) -> openpyxl.worksheet.worksheet.Worksheet | None:
    """Return the first sheet whose name contains any of the candidate strings (case-insensitive)."""
    for sheet_name in wb.sheetnames:
        for candidate in candidates:
            if candidate.lower() in sheet_name.lower():
                return wb[sheet_name]
    return None


def cell_str(cell) -> str:
    """Return stripped lowercase string value of a cell, or empty string."""
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip().lower()


def cell_float(cell) -> float | None:
    """Return float value of a cell, or None."""
    if cell is None or cell.value is None:
        return None
    try:
        return float(cell.value)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Reference consumption lookup
#
# Both file types store annual costs (£/year, ex-VAT) rather than unit rates.
# We derive rates using:
#   unit_rate (p/kWh, ex-VAT) = (annual_total - nil_annual) * 100 / ref_kwh
#   standing_charge (p/day, ex-VAT) = nil_annual * 100 / 365
#
# The reference kWh is encoded in the column code header, e.g.:
#   "ElecSingle_Other_3100kWh"  → 3100 kWh electricity
#   "Gas_Other_12000kWh"        → 12000 kWh gas
#   "ElecSingle_Other_Benchmark" → look up in BENCHMARK_KWH below
#
# Ofgem updated reference consumption levels from Q1 2026 onwards.
# These values are taken from the published Annex 9 technical notes.
# ---------------------------------------------------------------------------

# Reference kWh for quarters using "Benchmark" column codes.
# Key: period_from date string, value: (elec_kwh, gas_kwh)
BENCHMARK_KWH = {
    "2026-01-01": (2900, 11100),  # Q1 2026: Ofgem updated benchmarks
    "2026-04-01": (2700, 11500),  # Q2 2026: further revision
}


def _parse_ref_kwh(col_code: str, period_from: str, fuel: str) -> float | None:
    """Extract the reference kWh from a column code like 'ElecSingle_Other_3100kWh'.
    Returns None if the value cannot be determined."""
    match = re.search(r'(\d+)kwh', col_code, re.IGNORECASE)
    if match:
        return float(match.group(1))
    if "benchmark" in col_code.lower():
        benchmarks = BENCHMARK_KWH.get(period_from)
        if benchmarks:
            return benchmarks[0] if fuel == "electricity" else benchmarks[1]
    return None


# ---------------------------------------------------------------------------
# Unified parser for both pre-levelised and Annex 9 files (all quarters)
#
# Both file types share the same layout in their "1a ..." output sheet:
#
#   Row ~9:  "Other Payment Method"  (section header — the DD-equivalent method)
#   Row ~11: column codes  e.g. "ElecSingle_Other_Nil", "ElecSingle_Other_3100kWh", ...
#   Row ~12: "Charge Restriction Region" header
#   Row ~13: sub-labels  "Nil kWh", "m (3,100 kWh)", ...
#   Rows 14+: one row per region — region name in col B, values in cols C–H
#
# Column layout (0-indexed, col A = 0):
#   col 1 (B): region name
#   col 2 (C): ElecSingle_Other_Nil     — elec nil annual cost (£/yr ex-VAT)
#   col 3 (D): ElecSingle_Other_NNNkWh  — elec total at reference consumption
#   col 4 (E): ElecMulti_Other_Nil      — multi-rate elec nil (we don't use this)
#   col 5 (F): ElecMulti_Other_NNNkWh
#   col 6 (G): Gas_Other_Nil            — gas nil annual cost
#   col 7 (H): Gas_Other_NNNkWh        — gas total at reference consumption
#
# All values are ex-VAT; we apply 5% VAT to produce the consumer-facing figures.
#
# Pre-levelised sheet name:  "1a Default tariff cap"
# Annex 9 sheet name:        "1a levelised DTC" or "1a Levelised DTC"
# ---------------------------------------------------------------------------

def parse_cap_sheet(path: str, label: str, period_from: str) -> dict:
    """Parse either a pre-levelised or Annex 9 cap model file.
    Returns a dict of {gsp_code: {electricity: {...}, gas: {...}}} with rates inc VAT."""
    wb = openpyxl.load_workbook(path, data_only=True)

    # Find the "1a" output sheet — name varies by file version
    sheet = find_sheet(wb, ["1a default tariff cap", "1a levelised dtc"])
    if sheet is None:
        # Broader fallback: any sheet starting with "1a"
        for name in wb.sheetnames:
            if name.lower().startswith("1a"):
                sheet = wb[name]
                break

    if sheet is None:
        print(f"  [WARNING] {label}: could not find '1a' output sheet. Sheets: {wb.sheetnames}")
        return {}

    # -----------------------------------------------------------------------
    # Step 1: Locate the "Other Payment Method" block.
    # Scan for the row containing the column codes ("ElecSingle_Other_Nil" etc.)
    # -----------------------------------------------------------------------
    col_code_row_idx = None
    for row in sheet.iter_rows(max_row=30):
        row_text = " ".join(cell_str(c) for c in row)
        if "elecsingle_other_nil" in row_text or "elecsingle_other" in row_text:
            col_code_row_idx = row[0].row
            break

    if col_code_row_idx is None:
        print(f"  [WARNING] {label}: could not find column code row in sheet '{sheet.title}'.")
        return {}

    # Read the column codes from that row
    col_codes = [cell_str(c) for c in sheet[col_code_row_idx]]

    # Find the columns we need by matching column code substrings
    col_region = None
    col_elec_nil = None
    col_elec_ref = None
    col_gas_nil  = None
    col_gas_ref  = None

    for i, code in enumerate(col_codes):
        if "elecsingle_other_nil" in code:
            col_elec_nil = i
        elif "elecsingle_other_" in code and col_elec_nil is not None and col_elec_ref is None:
            col_elec_ref = i
        elif "gas_other_nil" in code:
            col_gas_nil = i
        elif "gas_other_" in code and col_gas_nil is not None and col_gas_ref is None:
            col_gas_ref = i

    # Find the region column from the next row (the "Charge Restriction Region" row)
    for row in sheet.iter_rows(min_row=col_code_row_idx, max_row=col_code_row_idx + 3):
        for cell in row:
            if "charge restriction region" in cell_str(cell):
                col_region = cell.column - 1  # 0-indexed
                break
        if col_region is not None:
            break

    # If still not found, assume it's column B (index 1)
    if col_region is None:
        col_region = 1

    missing = [name for name, col in [
        ("elec nil", col_elec_nil), ("elec ref", col_elec_ref),
        ("gas nil", col_gas_nil), ("gas ref", col_gas_ref),
    ] if col is None]

    if missing:
        print(f"  [WARNING] {label}: could not locate column codes: {missing}")
        print(f"  Column codes found: {[c for c in col_codes if c]}")
        return {}

    # Extract reference kWh from column codes
    elec_ref_code = col_codes[col_elec_ref] if col_elec_ref is not None else ""
    gas_ref_code  = col_codes[col_gas_ref]  if col_gas_ref  is not None else ""
    elec_ref_kwh  = _parse_ref_kwh(elec_ref_code, period_from, "electricity")
    gas_ref_kwh   = _parse_ref_kwh(gas_ref_code,  period_from, "gas")

    if elec_ref_kwh is None or gas_ref_kwh is None:
        print(f"  [WARNING] {label}: could not determine reference kWh from column codes "
              f"'{elec_ref_code}', '{gas_ref_code}'")
        return {}

    # -----------------------------------------------------------------------
    # Step 2: Read data rows — scan from below the column code row until we
    # stop finding region names.
    # -----------------------------------------------------------------------
    results = {}
    for row in sheet.iter_rows(min_row=col_code_row_idx + 1):
        if col_region >= len(row):
            continue
        region_name = cell_str(row[col_region])
        if not region_name:
            continue

        # Stop if we've moved past the Other section into Standard Credit / PPM
        if region_name in ("standard credit", "ppm", "prepayment"):
            break

        gsp = REGION_NAME_TO_GSP.get(region_name)
        if gsp is None:
            # Skip summary/average rows silently; warn on anything else
            if region_name not in ("charge restriction region", "gb average",
                                   "gb average, inc vat (at 5%)", "england",
                                   "scotland", "wales", "nil kwh"):
                print(f"  [WARNING] {label}: unrecognised region '{region_name}' — skipping")
            continue

        elec_nil = cell_float(row[col_elec_nil])
        elec_tot = cell_float(row[col_elec_ref])
        gas_nil  = cell_float(row[col_gas_nil])
        gas_tot  = cell_float(row[col_gas_ref])

        if not all(v is not None and v > 0 for v in [elec_nil, elec_tot, gas_nil, gas_tot]):
            continue

        # Derive rates from annual cost figures (ex-VAT, £/yr)
        # Apply 5% VAT to get consumer-facing inc-VAT figures
        vat = 1.05
        elec_ur_inc_vat = round(((elec_tot - elec_nil) * 100 / elec_ref_kwh) * vat, 4)
        elec_sc_inc_vat = round((elec_nil * 100 / 365) * vat, 4)
        gas_ur_inc_vat  = round(((gas_tot - gas_nil) * 100 / gas_ref_kwh) * vat, 4)
        gas_sc_inc_vat  = round((gas_nil * 100 / 365) * vat, 4)

        results[gsp] = {
            "electricity": {
                "unit_rate_inc_vat":       elec_ur_inc_vat,
                "standing_charge_inc_vat": elec_sc_inc_vat,
            },
            "gas": {
                "unit_rate_inc_vat":       gas_ur_inc_vat,
                "standing_charge_inc_vat": gas_sc_inc_vat,
            },
        }

    print(f"  [parsed] {label}: {len(results)}/14 regions")
    return results


# Keep old names as aliases so the main() call below still works
def parse_annex9(path: str, label: str, period_from: str = "") -> dict:
    return parse_cap_sheet(path, label, period_from)

def parse_pre_levelised(path: str, label: str, period_from: str = "") -> dict:
    return parse_cap_sheet(path, label, period_from)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_json():
    output = {
        "_meta": {
            "description": "Ofgem energy price cap rates by region for direct debit customers. Unit rates in p/kWh inc VAT. Standing charges in p/day inc VAT.",
            "source": "https://www.ofgem.gov.uk/energy-regulation/domestic-and-non-domestic/energy-pricing-rules/energy-price-cap/energy-price-cap-default-tariff-levels",
            "payment_method": "direct_debit",
            "generated": date.today().isoformat(),
            "regions": REGIONS,
        },
        "quarters": []
    }

    all_quarters = (
        [("pre_levelised", q) for q in PRE_LEVELISED_QUARTERS] +
        [("annex9",        q) for q in ANNEX9_QUARTERS]
    )

    failed = []

    for file_type, quarter in all_quarters:
        print(f"\n{quarter['label']} ({quarter['period_from']} – {quarter['period_to']})")
        try:
            path = download_file(quarter["url"], quarter["label"])
            if file_type == "annex9":
                rates = parse_annex9(path, quarter["label"], quarter["period_from"])
            else:
                rates = parse_pre_levelised(path, quarter["label"], quarter["period_from"])

            if rates:
                output["quarters"].append({
                    "period_from":  quarter["period_from"],
                    "period_to":    quarter["period_to"],
                    "label":        quarter["label"],
                    "file_type":    file_type,
                    "regions":      rates,
                })
            else:
                print(f"  [FAILED] No data extracted for {quarter['label']}")
                failed.append(quarter["label"])

        except Exception as e:
            print(f"  [ERROR] {quarter['label']}: {e}")
            failed.append(quarter["label"])

    # Sort by period
    output["quarters"].sort(key=lambda q: q["period_from"])

    out_path = "ofgem-price-cap.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2)

    print(f"\n{'='*60}")
    print(f"Done. Wrote {len(output['quarters'])} quarters to {out_path}")

    if failed:
        print(f"\nWARNING: {len(failed)} quarter(s) failed or produced no data:")
        for label in failed:
            print(f"  - {label}")
        print("\nFor failed quarters, open the XLSX file manually and check:")
        print("  1. Which sheet contains the consumer-facing DD rates table")
        print("  2. Whether region names match REGION_NAME_TO_GSP in this script")
        print("  3. Update the script and re-run (cached files won't re-download)")
    else:
        print("\nAll quarters parsed successfully.")

    print(f"\nNext quarterly update:")
    print(f"  1. Visit the Ofgem page (URL in script header)")
    print(f"  2. Add the new Annex 9 URL to ANNEX9_QUARTERS")
    print(f"  3. Run: python build_ofgem_data.py")


if __name__ == "__main__":
    build_json()
